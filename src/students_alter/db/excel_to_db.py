from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

from src.students_alter.db import implementing_alter_abc as imp


class AddStudents:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.is_file():
            raise ValueError("""Incorrect path!
Make to provide a full path.""")
        if self.file_path.suffix != ".xlsx":
            raise ValueError("""Incorrect file type!
Make sure the extension is 'xlsx'""")

        self.workbook = load_workbook(self.file_path)
        self.sheet = self.workbook.active

        self.new_class = imp.NewClass(self.sheet["E1"].value)
        self.new_class.alter()

    def alter(self):
        self.sheet["D1"] = "added"

        for row, student_ex in enumerate(self.sheet.iter_rows(min_row=2, values_only=True)):
            try:
                student_db = imp.NewStudent(
                    student_ex[0],
                    student_ex[1],
                    student_ex[2],
                    self.sheet["E1"].value
                )
            except ValueError:
                self.sheet[f"D{row + 2}"] = "no"
            else:
                student_db.alter()
                self.sheet[f"D{row + 2}"] = "yes"

        red_background = PatternFill(bgColor=colors.Color(rgb="f70523"))
        diff_style = DifferentialStyle(fill=red_background)
        rule = Rule(type="expression", dxf=diff_style)
        rule.formula = ['$D1="no"']
        self.sheet.conditional_formatting.add("A1:E100", rule)
        self.workbook.save(self.file_path)


if __name__ == "__main__":
    path = r"C:\Users\Mateusz\Downloads\criterion_B_rot_draft.xlsx"
    new_student = AddStudents(path)
